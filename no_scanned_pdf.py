import fitz  # PyMuPDF for text extraction
import pdfplumber  # Extract tables
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def extract_text(pdf_path):
    """Extracts text from a PDF and ensures each page starts on a new row."""
    extracted_text = []
    
    with fitz.open(pdf_path) as doc:
        for page_num, page in enumerate(doc, start=1):
            text = page.get_text("text").strip()
            if text:
                extracted_text.append((page_num, text))
    
    return extracted_text

def extract_tables(pdf_path):
    """Extracts tables from a PDF and returns them as DataFrames."""
    tables = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            extracted_tables = page.extract_tables()
            for table in extracted_tables:
                df = pd.DataFrame(table[1:], columns=table[0])  # Use first row as headers
                df.insert(0, "Page", page_num)  # Add page number column
                tables.append(df)
    
    return tables

def save_to_excel(text_data, tables, output_path="output.xlsx"):
    """Saves extracted text and tables into an Excel file."""
    wb = Workbook()
    
    # Save text data
    ws_text = wb.active
    ws_text.title = "Extracted Text"

    # Write headers
    ws_text.append(["Page", "Extracted Text"])
    ws_text["A1"].font = ws_text["B1"].font = Font(bold=True)

    row = 2  # Start writing from row 2

    for page_num, text in text_data:
        # Split text into chunks to avoid cell limits
        max_chars = 32000
        chunks = [text[i:i+max_chars] for i in range(0, len(text), max_chars)]
        
        for chunk in chunks:
            ws_text.append([page_num, chunk])
            row += 1  # Move to next row

    # Adjust column width and wrap text
    for col in ws_text.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)

    # Save tables in separate sheets
    for idx, table in enumerate(tables, start=1):
        ws_table = wb.create_sheet(title=f"Table {idx}")
        
        # Write headers
        for col_idx, column in enumerate(table.columns, start=1):
            cell = ws_table.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write table data
        for row_idx, row_data in enumerate(table.itertuples(index=False, name=None), start=2):
            ws_table.append(row_data)

    # Save file
    wb.save(output_path)
    print(f"✅ Data successfully saved to {output_path}")

# 📌 Run the script
pdf_file = "C:\\Users\\hp\\Downloads\\Contracts\\Contracts\\Centara Muscat.pdf" # Change this to your PDF file path
text_data = extract_text(pdf_file)
tables = extract_tables(pdf_file)
save_to_excel(text_data, tables, output_path="extract.xlsx")
